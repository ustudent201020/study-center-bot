import asyncio
import json
import random
import string
import os
import asyncpg
from aiogram import types
from aiogram.types import ParseMode, InputFile, ReplyKeyboardMarkup, KeyboardButton
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command, CommandStart
from openpyxl import Workbook

from data.config import CHANNELS, ADMINS
from keyboards.default.all import number, menu
from keyboards.inline.all import check_button, invite_button
from loader import bot, dp, db
from states.rekStates import Number, DelUser
from utils.misc import subscription
from keyboards.default.all import menu
from keyboards.default.rekKeyboards import admin_key, main_menu
from keyboards.default.rekKeyboards import back
from states.rekStates import RekData


@dp.message_handler(text='fix', user_id=ADMINS)
async def update_scoreee(message: types.Message):
    await message.answer('id va balni kiriting')
    await Number.add_user.set()


@dp.message_handler(state=Number.add_user)
async def fixx(message: types.Message, state: FSMContext):
    user_text = message.text.split(',')
    await db.update_user_score(score=int(user_text[0]), telegram_id=int(user_text[1]))
    await message.answer('bo`ldi')
    await state.finish()


@dp.message_handler(commands=['del'])
async def delete_user(message: types.Message, state: FSMContext):
    await message.answer('Id ni kiriting')
    await DelUser.user.set()


@dp.message_handler(state=DelUser.user)
async def delete(message: types.Message, state: FSMContext):
    await db.delete_users(telegram_id=int(f'{message.text}'))
    await message.answer('O"chirildi')
    await state.finish()


@dp.message_handler(commands=['upscore'])
async def delete_user(message: types.Message, state: FSMContext):
    await db.update_user_score(score=0, telegram_id=message.from_user.id)
    await message.answer('0')


@dp.message_handler(CommandStart())
async def show_channels(message: types.Message, state: FSMContext):
    args = message.get_args()
    if_old = await db.select_user(telegram_id=message.from_user.id)
    elements = await db.get_elements()
    photo = ''
    gifts = ''
    for element in elements:
        photo += f"{element['photo']}"
        gifts += f"{element['gifts']}"

    if args and not if_old:
        try:
            user = await db.add_user(telegram_id=message.from_user.id,
                                     full_name=message.from_user.full_name,
                                     username=message.from_user.username
                                     )
        except asyncpg.exceptions.UniqueViolationError:
            user = await db.select_user(telegram_id=message.from_user.id)
        await db.update_user_oldd(oldd='old', telegram_id=message.from_user.id)
        await db.update_user_args(user_args=f'{args}', telegram_id=message.from_user.id)

        status = True
        all = await db.select_chanel()
        chanels = []
        url = []
        channel_names = []
        for i in all:
            chanels.append(i['chanelll'])
            url.append(i['url'])
            channel_names.append(i['channel_name'])

        for channel in chanels:
            status *= await subscription.check(user_id=message.from_user.id,
                                               channel=f'{channel}')
        if status:
            user = await db.select_user(telegram_id=message.from_user.id)
            if user[3] is None or user[4] == 0:
                # result = f"<b>Tabriklaymiz ✅, Siz muvaffaqqiyatli ro`yxatdan o`tdingiz!</b>"
                # await message.answer(result, disable_web_page_preview=True)
                # await message.answer_photo(photo)
                await message.answer(text=f"{gifts}")
                await message.answer(
                    '<b>📲 Raqamni yuborish"</b> tugmasini bosgan holda raqamingizni yuboring!',
                    reply_markup=number,
                    disable_web_page_preview=True
                )
                await Number.number.set()
        else:
            button = types.InlineKeyboardMarkup(row_width=1, )
            counter = 0
            for i in url:
                button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
                counter += 1
            button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

            await message.answer(
                '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
                ' tugmasini bosing.',
                reply_markup=button,
                disable_web_page_preview=True)

    elif not args and not if_old:
        try:
            user = await db.add_user(telegram_id=message.from_user.id,
                                     full_name=message.from_user.full_name,
                                     username=message.from_user.username

                                     )
        except asyncpg.exceptions.UniqueViolationError:
            user = await db.select_user(telegram_id=message.from_user.id)
        status = True
        all = await db.select_chanel()
        chanels = []
        url = []
        channel_names = []
        for i in all:
            chanels.append(i['chanelll'])
            url.append(i['url'])
            channel_names.append(i['channel_name'])

        for channel in chanels:
            status *= await subscription.check(user_id=message.from_user.id,
                                               channel=f'{channel}')
        if status:
            user = await db.select_user(telegram_id=message.from_user.id)
            if user[3] is None or user[4] == 0:
                # result = f"<b>Tabriklaymiz ✅, Siz muvaffaqqiyatli ro`yxatdan o`tdingiz!</b>"
                # await message.answer_photo(photo)
                await message.answer(text=f"{gifts}")
                await message.answer(
                    '<b>📲 Raqamni yuborish" tugmasini bosgan holda raqamingizni yuboring!</b>',
                    reply_markup=number,
                    disable_web_page_preview=True
                )
                await Number.number.set()
            else:
                await message.answer("<b>Quyidagi menudan kerakli bo`limni tanlang 👇</b>",
                                     reply_markup=main_menu, disable_web_page_preview=True)
        else:
            button = types.InlineKeyboardMarkup(row_width=1, )
            counter = 0
            for i in url:
                button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
                counter += 1
            button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

            await message.answer(
                '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
                ' tugmasini bosing.',
                reply_markup=button,
                disable_web_page_preview=True)
    else:
        try:
            user = await db.add_user(telegram_id=message.from_user.id,
                                     full_name=message.from_user.full_name,
                                     username=message.from_user.username
                                     )
        except asyncpg.exceptions.UniqueViolationError:
            user = await db.select_user(telegram_id=message.from_user.id)
        status = True
        all = await db.select_chanel()
        chanels = []
        url = []
        channel_names = []
        for i in all:
            chanels.append(i['chanelll'])
            url.append(i['url'])
            channel_names.append(i['channel_name'])

        for channel in chanels:
            status *= await subscription.check(user_id=message.from_user.id,
                                               channel=f'{channel}')
        if status:
            user = await db.select_user(telegram_id=message.from_user.id)
            if user[3] is None or user[4] == 0:
                # result = f"<b>Tabriklaymiz ✅, Siz muvaffaqqiyatli ro`yxatdan o`tdingiz!</b>"
                # await message.answer_photo(photo)
                await message.answer(text=f"{gifts}")
                await message.answer(
                    '<b>📲 Raqamni yuborish" tugmasini bosgan holda raqamingizni yuboring!</b>',
                    reply_markup=number,
                    disable_web_page_preview=True
                )
                await Number.number.set()
            else:
                await message.answer("<b>Quyidagi menudan kerakli bo`limni tanlang 👇</b>",
                                     reply_markup=main_menu, disable_web_page_preview=True)
        else:
            button = types.InlineKeyboardMarkup(row_width=1, )
            counter = 0
            for i in url:
                button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
                counter += 1
            button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

            await message.answer(
                '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
                ' tugmasini bosing.',
                reply_markup=button,
                disable_web_page_preview=True)


@dp.callback_query_handler(text="check_subs")
async def checker(call: types.CallbackQuery, state: FSMContext):
    # await call.answer()
    result = str()
    result2 = str()
    status = True
    all = await db.select_chanel()
    chanels = []
    url = []
    channel_names = []
    for i in all:
        chanels.append(i['chanelll'])
        url.append(i['url'])
        channel_names.append(i['channel_name'])

    elements = await db.get_elements()
    photo = ''
    gifts = ''
    for element in elements:
        photo += f"{element['photo']}"
        gifts += f"{element['gifts']}"

    for channel in chanels:
        status *= await subscription.check(user_id=call.from_user.id,
                                           channel=f'{channel}')
    if status:
        try:
            await call.message.delete()
        except Exception as e:
            pass
        if_old = await db.select_user(telegram_id=call.from_user.id)
        if if_old[3] is None or if_old[4] == 0:
            # result += f"<b>Tabriklaymiz ✅, Siz muvaffaqqiyatli ro`yxatdan o`tdingiz!</b>"
            # await call.message.answer(result, disable_web_page_preview=True)
            # await call.message.answer_photo(photo=photo)
            await call.message.answer(text=f"{gifts}")
            await call.message.answer(
                '<b>📲 Raqamni yuborish" tugmasini bosgan holda raqamingizni yuboring!</b>',
                reply_markup=number,
                disable_web_page_preview=True
            )
            await Number.number.set()
            await bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
        else:
            await call.message.answer("<b>Quyidagi menudan kerakli bo`limni tanlang 👇</b>",
                                      reply_markup=main_menu, disable_web_page_preview=True)

    else:
        button = types.InlineKeyboardMarkup(row_width=1, )
        counter = 0
        for i in url:
            button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
            counter += 1
        button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

        try:
            await call.message.edit_text(f'❌ Kanalga aʼzo boʼlmadingiz!Botdan'
                                         f' toʼliq foydalanish uchun koʼrsatilgan barcha kanallarga aʼzo boʼling!\nKeyin <b>“A’zo bo’ldim”</b> tugmasini bosing.',
                                         reply_markup=button,
                                         disable_web_page_preview=True)
        except Exception as e:
            pass


@dp.message_handler(state=Number.number, content_types=types.ContentType.CONTACT)
async def phone_number(message: types.Message, state: FSMContext):
    elements = await db.get_elements()
    scoree = 0
    for element in elements:
        scoree += element['limit_score']

    user = await db.select_user(telegram_id=message.from_user.id)
    numberr = f'{message.contact.phone_number}'
    if numberr.startswith("+998") or numberr.startswith("998"):
        if user[3] is None or user[4] == 0:
            try:
                user_telephone_num = await db.update_user_phone(phone=message.contact.phone_number,
                                                                telegram_id=message.from_user.id)
            except:
                pass
            if user[4] == 0 or user[4] is None:
                user = await db.update_user_score(score=scoree, telegram_id=message.from_user.id)
            try:
                args = await db.select_user(telegram_id=message.from_user.id)
                args_user = await db.select_user(telegram_id=int(args[7]))

                update_score = int(args_user[4]) + scoree
                await db.update_user_score(score=update_score, telegram_id=int(args[7]))
                await bot.send_message(chat_id=int(args[7]),
                                       text=f"👤 Yangi ishtirokchi qo`shildi\n"
                                            f"🎗 Sizning balingiz {update_score},"
                                            f" ko`proq do`stlaringizni taklif qiling!")
            except Exception as e:
                pass
            # await message.answer(f"<b>🎉 Tabriklaymiz ✅, Siz boshlang`ich {scoree} balga ega bo`ldingiz!</b>",
            # disable_web_page_preview=True)
            await message.answer("<b>🎉Tabriklaymiz, siz tanlov ishtirokchisiga aylandingiz!</b>\n\n"
                                 "Quyidagi menudan kerakli bo`limni tanlang 👇",
                                 reply_markup=main_menu, disable_web_page_preview=True)
            await state.finish()
        else:
            await message.answer("<b>Quyidagi menudan kerakli bo`limni tanlang 👇</b>",
                                 reply_markup=main_menu, disable_web_page_preview=True)
            await state.finish()
    else:
        await message.answer('Kechirasiz Faqat O`zbekiston raqamlarini qabul qilamiz',
                             reply_markup=types.ReplyKeyboardRemove())
        await state.finish()


@dp.message_handler(text='fix', user_id=ADMINS)
async def update_scoreee(message: types.Message):
    await message.answer('id va balni kiriting')
    await DelUser.fix.set()


@dp.message_handler(state=DelUser.fix)
async def fixx(message: types.Message, state: FSMContext):
    user_text = message.text.split(',')
    await db.update_user_score(score=int(user_text[0]), telegram_id=int(user_text[1]))
    await message.answer('bo`ldi')
    await state.finish()


@dp.message_handler(text='🎁 Tanlovda ishtirok etish')
async def tanlov(message: types.Message):
    elements = await db.get_elements()
    photo = ''
    txt = ''
    for element in elements:
        photo += f"{element['photo']}"
        txt += f"{element['game_text']}"

    status = True
    all = await db.select_chanel()
    chanels = []
    url = []
    channel_names = []
    for i in all:
        chanels.append(i['chanelll'])
        url.append(i['url'])
        channel_names.append(i['channel_name'])

    for channel in chanels:
        status *= await subscription.check(user_id=message.from_user.id,
                                           channel=f'{channel}')
    if status:
        txt += f'\n\nhttp://t.me/goenglishuzbot?start={message.from_user.id}'
        await message.answer_photo(photo=photo,
                                   caption=txt,
                                   parse_mode='HTML'
                                   )
        await message.answer(
            '👆 Yuqoridagi sizning referal <b>link/havolangiz</b>. Uni koʼproq tanishlaringizga ulashing. Omad!')

    else:
        button = types.InlineKeyboardMarkup(row_width=1, )
        counter = 0
        for i in url:
            button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
            counter += 1
        button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

        await message.answer(
            '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
            ' tugmasini bosing.',
            reply_markup=button,
            disable_web_page_preview=True)


@dp.message_handler(text='🎁 Sovg`alar')
async def my_score(message: types.Message):
    elements = await db.get_elements()
    photo = ''
    txt = ''
    for element in elements:
        photo += f"{element['photo']}"
        txt += f"{element['gifts']}"

    status = True
    all = await db.select_chanel()
    chanels = []
    url = []
    channel_names = []
    for i in all:
        chanels.append(i['chanelll'])
        url.append(i['url'])
        channel_names.append(i['channel_name'])

    for channel in chanels:
        status *= await subscription.check(user_id=message.from_user.id,
                                           channel=f'{channel}')
    if status:
        # await message.answer_photo(photo)
        await message.answer(text=txt, disable_web_page_preview=True)

    else:
        button = types.InlineKeyboardMarkup(row_width=1, )
        counter = 0
        for i in url:
            button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
            counter += 1
        button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

        await message.answer(
            '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
            ' tugmasini bosing.',
            reply_markup=button,
            disable_web_page_preview=True)


@dp.message_handler(text='👤 Ballarim')
async def my_score(message: types.Message):
    status = True
    all = await db.select_chanel()
    chanels = []
    url = []
    channel_names = []
    for i in all:
        chanels.append(i['chanelll'])
        url.append(i['url'])
        channel_names.append(i['channel_name'])

    for channel in chanels:
        status *= await subscription.check(user_id=message.from_user.id,
                                           channel=f'{channel}')
    if status:
        score = await db.select_user(telegram_id=message.from_user.id)
        await message.answer(f'📌Sizda {score[4]} ball mavjud.')
    else:
        button = types.InlineKeyboardMarkup(row_width=1, )
        counter = 0
        for i in url:
            button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
            counter += 1
        button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

        await message.answer(
            '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
            ' tugmasini bosing.',
            reply_markup=button,
            disable_web_page_preview=True)


# @dp.message_handler(text='🧑🏻‍💻 Админ')
# async def admin(message: types.Message):
#     status = True
#     for channel in CHANNELS:
#         status *= await subscription.check(user_id=message.from_user.id,
#                                            channel=channel)
#     if status:
#         await message.answer(f'@Dilshodbek_Zubaydov1')
#     else:
#         await message.answer(f'Tanlovda ishtirok etish учун қуйидаги 6 каналга аъзо бўлинг. '
#                              f'Кейин "Аъзо бўлдим" тугмасини босинг',
#                              reply_markup=check_button,
#                              disable_web_page_preview=True)

@dp.message_handler(text='Statistika 📊')
async def show_users(message: types.Message):
    a = await db.count_users()
    await message.answer(f'<b>🔷 Жами обуначилар: {a} та</b>')


@dp.message_handler(text='📊 Reyting')
async def score(message: types.Message):
    status = True
    all = await db.select_chanel()
    chanels = []
    url = []
    channel_names = []
    for i in all:
        chanels.append(i['chanelll'])
        url.append(i['url'])
        channel_names.append(i['channel_name'])

    for channel in chanels:
        status *= await subscription.check(user_id=message.from_user.id,
                                           channel=f'{channel}')
    if status:
        ball = await db.select_user(telegram_id=message.from_user.id)
        counter = 1
        text = '<b>📊 Tanlovimizda eng koʼp doʼstini taklif qilib, yuqori ball toʼplaganlar reytingi:</b>\n\n'
        elements = await db.get_elements()
        winners = 21
        list_all_score = await db.select_top_users_list()
        user_score = 0
        user_order = 0
        # print(list_all_score)
        for i in list_all_score:
            # if user_order != 0:
            user_order += 1
            if i[6] == message.from_user.id:
                user_score += i[4]
                break

        # for i in elements:
        #     winners += int(i["winners"])
        # top = await db.select_top_users(lim_win=winners)
        for i in list_all_score:
            text += f"<b>🏅{counter}-o'rin</b> : {i[1]} - {i[4]} ta\n"
            counter += 1
            if counter == winners:
                break
        if counter:
            text += f'<b>...\n{user_order}-o`rin: {message.from_user.full_name}</b> - {user_score} ta\n\n✅ Sizda <b>{ball[4]} ball</b> mavjud.\n\n' \
                    f'Koʼproq doʼstlaringizni taklif qilib, ballingizni oshiring!\n\n' \
                    f'👤 Sizning referal link/havolangiz:\n ' \
                    f'https://t.me/Barakali_tanlov_bot?start={message.from_user.id} \n' \
                    f'<b>Uni koʼproq tanishlaringizga ulashing. Omad!</b>   '
            await message.answer(text=text, disable_web_page_preview=True)
    else:
        button = types.InlineKeyboardMarkup(row_width=1, )
        counter = 0
        for i in url:
            button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
            counter += 1
        button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

        await message.answer(
            '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
            ' tugmasini bosing.',
            reply_markup=button,
            disable_web_page_preview=True)


@dp.message_handler(text='💡 Shartlar')
async def help(message: types.Message):
    status = True
    all = await db.select_chanel()
    chanels = []
    url = []
    channel_names = []
    for i in all:
        chanels.append(i['chanelll'])
        url.append(i['url'])
        channel_names.append(i['channel_name'])

    for channel in chanels:
        status *= await subscription.check(user_id=message.from_user.id,
                                           channel=f'{channel}')
    if status:
        elements = await db.get_elements()
        photo = ''
        shartlar = ''
        for element in elements:
            photo += f"{element['photo']}"
            shartlar += f"{element['shartlar']}"
        # await message.answer_photo(caption=shartlar, photo=photo,parse_mode='HTML')
        await message.answer(text=shartlar, disable_web_page_preview=True)

    else:
        button = types.InlineKeyboardMarkup(row_width=1, )
        counter = 0
        for i in url:
            button.add(types.InlineKeyboardButton(f"{channel_names[counter]}", url=f'https://t.me/{i}'))
            counter += 1
        button.add(types.InlineKeyboardButton(text="A’zo bo’ldim", callback_data="check_subs"))

        await message.answer(
            '✅Tanlovda ishtirok etish uchun quyidagi kanallarga a’zo bo’ling.\nKeyin <b>“A’zo bo’ldim”</b>'
            ' tugmasini bosing.',
            reply_markup=button,
            disable_web_page_preview=True)


@dp.message_handler(Command('jsonFile'))
async def jsonnn(message: types.Message):
    user_list = []
    userss = await db.select_all_users()
    for user in userss:
        user_dict = {}
        user_dict['full_name'] = user[1]
        user_dict['username'] = user[2]
        user_dict['phone'] = user[3]
        user_dict['score'] = user[4]
        user_dict['tg_id'] = user[6]
        user_list.append(user_dict)
        await asyncio.sleep(0.05)
    with open("users.json", "w") as outfile:
        json.dump(user_list, outfile)
    document = open('users.json')
    await bot.send_document(message.from_user.id, document=document)


@dp.message_handler(text="Excel File")
async def marathon(message: types.Message):
    elements = await db.get_elements()
    winners_ball = 1
    for i in elements:
        winners_ball += int(i["winners"])

    wb = Workbook()
    ws = wb.active
    ws['A1'] = "№"
    ws['B1'] = 'NAME'
    ws['C1'] = "BAll"

    # Rows can also be appended
    userss = await db.select_top_users_list()
    counter = 0
    for user in userss:
        if user[4] <= winners_ball:
            continue
        else:
            counter += 1
            ws.append([f"{counter}-o'rin", f"{user[1]}", f"{user[4]}-ball"])

    # Python types will automatically be converted
    # import datetime
    # ws['A2'] = datetime.datetime.now()
    # for i in ws['A1']:
    #     print(i.value)
    n = random.choices(string.ascii_lowercase, k=2)
    c = random.choices(string.ascii_lowercase, k=2)
    m = random.choices(string.ascii_lowercase, k=2)

    wb.save(f"{n}-{m}.xlsx")
    file = InputFile(path_or_bytesio=f'{n}-{m}.xlsx')
    await message.answer_document(document=file)
    os.remove(f"{n}-{m}.xlsx")


@dp.message_handler(text="G'oliblar haqida ma'lumot")
async def scoree(message: types.Message):
    ball = await db.select_user(telegram_id=message.from_user.id)
    counter = 1
    text = '<b>📊 Tanlovimizda eng koʼp doʼstini taklif qilib, yuqori ball toʼplaganlar reytingi:</b>\n\n'
    elements = await db.get_elements()
    winners = 21
    list_all_score = await db.select_top_users_list()
    user_score = 0
    user_order = 0
    # print(list_all_score)
    for i in list_all_score:
        # if user_order != 0:
        user_order += 1
        if i[6] == message.from_user.id:
            user_score += i[4]
            break

    # for i in elements:
    #     winners += int(i["winners"])
    # top = await db.select_top_users(lim_win=winners)
    for i in list_all_score:
        text += f"🏅{counter}-o'rin    <a href='tg://user?id={i[6]}'> {i[1]} </a> • {i[4]} ball," \
                f" username: @{i[2]}, tg_id: {i[6]} phone: {i[3]}\n"
        counter += 1
        if counter == winners:
            break
    if counter:
        # text += f'<b>...\n{user_order}-o`rin: {message.from_user.full_name}</b> - {user_score} ta\n\n✅ Sizda <b>{ball[4]} ball</b> mavjud.\n\n' \
        #         f'Koʼproq doʼstlaringizni taklif qilib, ballingizni oshiring!\n\n' \
        #         f'👤 Sizning referal link/havolangiz:\n ' \
        #         f'https://t.me/Barakali_tanlov_bot?start={message.from_user.id} \n' \
        #         f'<b>Uni koʼproq tanishlaringizga ulashing. Omad!</b>   '
        await message.answer(text=text, disable_web_page_preview=True)


@dp.message_handler(Command('read_file'))
async def json_reader(message: types.Message):
    f = open('users.json', 'r')
    data = json.loads(f.read())
    for user in data:
        # print(user['tg_id'])
        try:
            user = await db.add_json_file_user(
                telegram_id=user['tg_id'],
                username=user['username'],
                full_name=user['full_name'],
                phone=user['phone'],
                score=user['score']
            )
        except Exception as e:
            print(e)
    f.close()


@dp.message_handler(commands=['dasturchi'])
async def i_2(message: types.Message):
    await message.answer(
        "🧑‍💻Dasturchi: <a href='http://t.me/calll_robot'>Ilyosbek</a> 🧑‍💻",
        disable_web_page_preview=True)
